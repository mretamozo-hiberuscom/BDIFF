"""Render a color-coded Excel workbook from a ComparisonResult using FindingView."""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from schema_comparator.compare.models import (
    ColumnMismatch,
    ComparisonResult,
    DiffEntry,
    MissingColumn,
    MissingProcedure,
    MissingTable,
    ProcedureMismatch,
)
from schema_comparator.report.attributes import MISSING_MARKER
from schema_comparator.report.presentation import present_finding
from schema_comparator.report.rows import grouped_rows_by_table

_FILL_MISSING_TABLE = PatternFill("solid", fgColor="FDE2E1")
_FILL_MISSING_COLUMN = PatternFill("solid", fgColor="FFF3CD")
_FILL_MISMATCH = PatternFill("solid", fgColor="D9EDF7")
_FILL_HEADER = PatternFill("solid", fgColor="E9ECEF")

_LEGEND: tuple[tuple[str, PatternFill], ...] = (
    ("Tabla faltante", _FILL_MISSING_TABLE),
    ("Columna faltante", _FILL_MISSING_COLUMN),
    ("Discrepancia de atributos", _FILL_MISMATCH),
)


def _fill_for(entry: DiffEntry) -> PatternFill:
    if isinstance(entry, (MissingTable, MissingProcedure)):
        return _FILL_MISSING_TABLE
    if isinstance(entry, MissingColumn):
        return _FILL_MISSING_COLUMN
    return _FILL_MISMATCH


def _write_findings_sheet(ws: Worksheet, result: ComparisonResult) -> None:
    header = [
        "Esquema",
        "Tabla",
        "Columna",
        "Tipo de diferencia",
        *result.compared_profiles,
    ]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = _FILL_HEADER

    for (schema_name, object_name), rows_of_entries in grouped_rows_by_table(result):
        for group in rows_of_entries:
            first = group[0]
            view = present_finding(group, result.compared_profiles)
            profile_cells = []
            for p in result.compared_profiles:
                val = view.cells_by_profile.get(p, "")
                if val == "Presente":
                    val = ""
                profile_cells.append(val)

            row = [
                view.schema_name,
                view.object_name,
                view.detail_name or "",
                view.finding_type,
                *profile_cells,
            ]
            ws.append(row)
            fill = _fill_for(first)
            for cell in ws[ws.max_row]:
                cell.fill = fill

    for col_index, column_cells in enumerate(ws.columns, start=1):
        longest = max((len(str(c.value)) for c in column_cells if c.value), default=0)
        ws.column_dimensions[get_column_letter(col_index)].width = min(
            max(longest + 2, 10), 60
        )


def _write_legend_sheet(ws: Worksheet) -> None:
    ws.append(["Color", "Significado"])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    for row_index, (label, fill) in enumerate(_LEGEND, start=2):
        color_cell = ws.cell(row=row_index, column=1)
        color_cell.fill = fill
        ws.cell(row=row_index, column=2, value=label)
    marker_row = len(_LEGEND) + 2
    ws.cell(
        row=marker_row,
        column=1,
        value=MISSING_MARKER,
    ).alignment = Alignment(horizontal="center")
    ws.cell(
        row=marker_row,
        column=2,
        value="No presente en ese perfil",
    )
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40


def export_excel(result: ComparisonResult) -> bytes:
    """Render `result` to a `.xlsx` workbook (bytes) using FindingView."""
    workbook = Workbook()
    findings_ws = workbook.active
    findings_ws.title = "Diferencias"
    _write_findings_sheet(findings_ws, result)

    legend_ws = workbook.create_sheet("Leyenda")
    _write_legend_sheet(legend_ws)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
