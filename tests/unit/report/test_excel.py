"""Unit tests for report.excel: export_excel (real openpyxl, no I/O)."""

import io

from openpyxl import load_workbook

from report.conftest import comparison_result_empty, comparison_result_with_findings

from schema_comparator.compare.models import ComparisonResult, MissingTable
from schema_comparator.report.excel import export_excel


def _load(xlsx_bytes: bytes):
    return load_workbook(io.BytesIO(xlsx_bytes))


def test_export_excel_creates_findings_and_legend_sheets() -> None:
    wb = _load(export_excel(comparison_result_with_findings()))

    assert wb.sheetnames == ["Diferencias", "Leyenda"]


def test_export_excel_header_row_lists_compared_profiles() -> None:
    wb = _load(export_excel(comparison_result_with_findings()))
    ws = wb["Diferencias"]

    header = [cell.value for cell in ws[1]]

    assert header == ["Esquema", "Tabla", "Columna", "Tipo de diferencia", "a", "b", "c"]


def test_export_excel_one_row_per_finding() -> None:
    result = comparison_result_with_findings()
    wb = _load(export_excel(result))
    ws = wb["Diferencias"]

    assert ws.max_row == 1 + len(result.entries)


def test_export_excel_missing_column_row_shows_present_side_attributes() -> None:
    wb = _load(export_excel(comparison_result_with_findings()))
    ws = wb["Diferencias"]

    row = next(
        r for r in ws.iter_rows(min_row=2, values_only=True) if r[2] == "notes"
    )
    # Esquema, Tabla, Columna, Tipo, a, b, c
    assert row[3] == "Columna faltante"
    assert row[4] == "varchar(255), NULL"
    assert row[5] == "varchar(255), NULL"
    assert row[6] == "\u274c"


def test_export_excel_empty_result_has_header_only() -> None:
    wb = _load(export_excel(comparison_result_empty()))
    ws = wb["Diferencias"]

    assert ws.max_row == 1


def test_export_excel_table_missing_from_multiple_profiles_is_a_single_row() -> None:
    result = ComparisonResult(
        compared_profiles=("autos", "decesos", "hogar", "vida"),
        entries=(
            MissingTable(schema_name="dbo", table_name="TC_Productos", missing_from_profile="decesos"),
            MissingTable(schema_name="dbo", table_name="TC_Productos", missing_from_profile="hogar"),
            MissingTable(schema_name="dbo", table_name="TC_Productos", missing_from_profile="vida"),
        ),
    )

    wb = _load(export_excel(result))
    ws = wb["Diferencias"]

    assert ws.max_row == 2  # header + one merged row
    row = next(ws.iter_rows(min_row=2, values_only=True))
    # Esquema, Tabla, Columna, Tipo, autos, decesos, hogar, vida
    # openpyxl round-trips an empty-string cell value as None.
    assert row == ("dbo", "TC_Productos", None, "Tabla faltante", None, "\u274c", "\u274c", "\u274c")


def test_export_excel_legend_sheet_lists_all_diff_categories() -> None:
    wb = _load(export_excel(comparison_result_with_findings()))
    ws = wb["Leyenda"]

    labels = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]

    assert labels == [
        "Tabla faltante",
        "Columna faltante",
        "Discrepancia de atributos",
        "No presente en ese perfil",
    ]
    assert ws.cell(row=ws.max_row, column=1).value == "\u274c"
